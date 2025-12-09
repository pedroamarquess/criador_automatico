from openpyxl import load_workbook
import pandas as pd
from separados_nomes import separar_nomes
import tkinter as tk
from tkinter import messagebox


class GratRepApp:

    def __init__(self):
        self.janela = tk.Tk()
        self.janela.config(background="white", pady=20, padx=20)
        self.janela.geometry("900x500")

        # Carrega o Excel base ao iniciar
        self.wb = load_workbook("novo_arquivo_base.xlsx")
        self.ws_info = self.wb["INFO"]

        # Pega nomes do CSV
        self.lista_graduacao, self.lista_nome = self.carregar_nomes()

        # Coloca os nomes no Excel
        self.preencher_militares()

        # Cria tela inicial
        self.widgets = []
        self.criar_formulario()

        self.janela.mainloop()


    def carregar_nomes(self):
        arquivo_csv_nomes = separar_nomes()
        df = pd.read_csv(arquivo_csv_nomes)
        return df["graduacao"].to_list(), df["nome_completo"].to_list()


    def preencher_militares(self):
        ws_qsgr = self.wb["QSGR"]
        numero_celula = 16

        for i in range(len(self.lista_nome)):
            ws_qsgr[f"B{numero_celula}"] = self.lista_graduacao[i]
            ws_qsgr[f"D{numero_celula}"] = self.lista_nome[i]
            numero_celula += 1


    def criar_formulario(self):
        """Cria os campos e botões da interface"""
        itens = [
            ("Insira o dia da saída e mês (DD MM):", tk.Entry),
            ("Insira o horário da saída (XXXX):", tk.Entry),
            ("BI de saída (Ex: BI Nr 206, de 05/11/2025):", tk.Entry),
            ("Dia e mês da chegada (DD MM):", tk.Entry),
            ("Horário da chegada (XXXX):", tk.Entry),
            ("BI de chegada (se igual ao da saída, digite igual):", tk.Entry),
            ("Cidade:", tk.Entry),
            ("Finalidade:", tk.Entry),
            ("Número do QSGR:", tk.Entry)
        ]

        self.inputs = []

        for idx, (texto, input_type) in enumerate(itens, start=1):
            label = tk.Label(text=texto, bg="white", fg="black", pady=10)
            label.grid(column=0, row=idx)
            self.widgets.append(label)

            entrada = input_type(self.janela)
            entrada.grid(column=1, row=idx)
            self.widgets.append(entrada)
            self.inputs.append(entrada)

        self.botao_enviar = tk.Button(
            text="Fazer Grat Rep",
            bg="white",
            fg="black",
            command=self.fazer_grat_rep
        )
        self.botao_enviar.grid(column=0, row=len(itens)+2, columnspan=2)
        self.widgets.append(self.botao_enviar)


    def fazer_grat_rep(self):
        try:
            dia_saida, mes_saida = self.inputs[0].get().split()
            hora_saida = self.inputs[1].get()
            bi_saida = self.inputs[2].get()

            dia_ch, mes_ch = self.inputs[3].get().split()
            hora_ch = self.inputs[4].get()

            bi_ch = bi_saida if self.inputs[5].get().lower() == "igual" else self.inputs[5].get()

            cidade = self.inputs[6].get()
            finalidade = self.inputs[7].get()
            numero_qsgr = self.inputs[8].get()

        except:
            messagebox.showerror("Erro", "Preencha todos os campos corretamente!")
            return

        # Preenche aba INFO
        ws = self.ws_info
        ws["C4"], ws["C5"], ws["C7"], ws["C9"] = dia_saida, mes_saida, hora_saida, bi_saida
        ws["C12"], ws["C13"], ws["C15"], ws["C17"] = dia_ch, mes_ch, hora_ch, bi_ch
        ws["C18"], ws["B22"], ws["C2"] = cidade, finalidade, numero_qsgr

        # Salva o arquivo modificado
        nome_arquivo = f"QSGR {numero_qsgr.upper()} {cidade.upper()} {finalidade.upper()}.xlsx"
        self.wb.save(nome_arquivo)

        self.botao_enviar.config(state=tk.DISABLED)

        self.janela.after(1000, self.refazer_tela)


    def refazer_tela(self):
        for widget in self.widgets:
            widget.grid_forget()

        messagebox.showinfo("Tudo certo!", "Arquivo criado com sucesso!")

        botao_novo = tk.Button(
            text="Criar outra Grat Rep",
            command=self.nova_solicitacao,
            bg="white", fg="black"
        )
        botao_novo.pack(expand=True)
        self.widgets = [botao_novo]


    def nova_solicitacao(self):
        for widget in self.widgets:
            widget.pack_forget()
        self.widgets.clear()
        self.criar_formulario()


if __name__ == "__main__":
    GratRepApp()
